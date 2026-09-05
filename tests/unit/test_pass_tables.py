from __future__ import annotations

from io import BytesIO
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from larp_bot.adapters.memory import MemoryEventRepository, MemoryRegistrationRepository, MemoryUserRepository
from larp_bot.adapters.yandex_disk.repository import PASS_TABLE_HEADERS, YandexDiskShowcaseRepository
from larp_bot.application.services import EventAdministrationService, OrderedMutationService, RegistrationCatalog
from larp_bot.domain.models import (
    AttendanceStatus,
    CharacterWishPayload,
    EmptyPayload,
    EnlistPayload,
    EventStatus,
    Operation,
    OrderedRegistrationCommand,
    PassDetails,
    Platform,
    Registration,
    TelegramUser,
    VkUser,
)
from larp_bot.domain.security import participant_key
from tests.conftest import MemoryDiskStore


def russian_pass() -> PassDetails:
    return PassDetails(
        surname_cyrillic="Иванов",
        name_cyrillic="Иван",
        patronym_cyrillic="-",
        foreigner=False,
        mobile_phone="+7 999 123-45-67",
        email="ivan@example.com",
    )


def foreign_pass() -> PassDetails:
    return PassDetails(
        surname_cyrillic="Ли",
        name_cyrillic="Анна",
        patronym_cyrillic="-",
        foreigner=True,
        surname_latin="Li",
        name_latin="Anna",
        patronym_latin="-",
        mobile_phone="+44 7700 900123",
        email="anna@example.com",
    )


@pytest.mark.asyncio
async def test_pass_table_uses_only_confirmed_pass_profiles_from_ydb(disk_store: MemoryDiskStore, event) -> None:
    secret = "participant-secret"
    users = MemoryUserRepository()
    player_profiles = [
        TelegramUser(
            tg_id=11,
            full_name="Иван Иванов",
            vk_url="https://vk.com/ivan",
            crossplay=False,
            larp_experience=True,
            needs_pass=True,
            pass_details=russian_pass(),
        ),
        VkUser(
            vk_id=22,
            full_name="Анна Ли",
            crossplay=True,
            larp_experience=False,
            needs_pass=True,
            pass_details=foreign_pass(),
        ),
        TelegramUser(
            tg_id=33,
            full_name="Ждущий Игрок",
            vk_url="https://vk.com/waiting",
            crossplay=False,
            larp_experience=False,
            needs_pass=True,
            pass_details=russian_pass().model_copy(update={"email": "waiting@example.com"}),
        ),
        TelegramUser(
            tg_id=44,
            full_name="Без Пропуска",
            vk_url="https://vk.com/no-pass",
            crossplay=False,
            larp_experience=False,
            needs_pass=False,
        ),
    ]
    for user in player_profiles:
        await users.save(user)

    registrations = []
    for user, status in zip(
        player_profiles,
        (
            AttendanceStatus.CONFIRMED,
            AttendanceStatus.CONFIRMED,
            AttendanceStatus.WAITING,
            AttendanceStatus.CONFIRMED,
        ),
        strict=True,
    ):
        platform = Platform.TELEGRAM if isinstance(user, TelegramUser) else Platform.VK
        uid = user.tg_id if isinstance(user, TelegramUser) else user.vk_id
        registrations.append(
            Registration(
                event_id=event.event_id,
                participant_key=participant_key(secret, platform, uid, event.event_id),
                display_name=user.full_name or "",
                wish_play="Без пожеланий",
                attendance_status=status,
            )
        )

    events = MemoryEventRepository([event])
    tables = MemoryRegistrationRepository(registrations)
    workbooks = YandexDiskShowcaseRepository(disk_store)
    catalog = RegistrationCatalog(events, tables, workbooks)
    service = EventAdministrationService(events, workbooks, catalog, users, secret)

    result = await service.create_pass_table(event.event_id)

    assert result.created is True
    assert result.row_count == 2
    assert result.public_url == "https://disk.example/public/1"
    stored_event = await events.get(event.event_id)
    assert stored_event is not None
    assert stored_event.pass_table_resource_path == f"disk:/larp-bot/passes/{event.event_id}.xlsx"
    assert stored_event.pass_table_public_url == result.public_url

    content = disk_store.files[stored_event.pass_table_resource_path]
    workbook = load_workbook(BytesIO(content))
    try:
        sheet = workbook.active
        assert sheet.title == "Лист1"
        assert tuple(cell.value for cell in sheet[1]) == PASS_TABLE_HEADERS
        assert all(cell.fill.fill_type == "solid" for cell in sheet[1])
        assert all(cell.fill.fgColor.type == "theme" and cell.fill.fgColor.theme == 5 for cell in sheet[1])
        assert sheet.column_dimensions["G"].width == sheet.column_dimensions["C"].width
        table_cells = (
            cell
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=len(PASS_TABLE_HEADERS))
            for cell in row
        )
        assert all(cell.font.name == "Calibri Light" and cell.font.size == 14 for cell in table_cells)
        rows = {
            sheet.cell(row, 1).value: tuple(sheet.cell(row, column).value for column in range(1, 10))
            for row in range(2, 4)
        }
        assert rows["Иванов"] == (
            "Иванов",
            "Иван",
            "-",
            "Нет",
            None,
            None,
            None,
            "+7 999 123-45-67",
            "ivan@example.com",
        )
        assert rows["Ли"] == (
            "Ли",
            "Анна",
            "-",
            "Да",
            "Li",
            "Anna",
            "-",
            "+44 7700 900123",
            "anna@example.com",
        )
    finally:
        workbook.close()

    repeated = await service.create_pass_table(event.event_id)
    assert repeated.public_url == result.public_url
    assert repeated.created is False
    assert repeated.row_count is None
    assert len(disk_store.files) == 1


@pytest.mark.asyncio
async def test_pass_table_tracks_confirmation_and_cancellation_without_changing_url(
    disk_store: MemoryDiskStore, event
) -> None:
    secret = "participant-secret"
    users = MemoryUserRepository()
    await users.save(
        TelegramUser(
            tg_id=11,
            full_name="Иван Иванов",
            vk_url="https://vk.com/ivan",
            crossplay=False,
            larp_experience=True,
            needs_pass=True,
            pass_details=russian_pass(),
        )
    )
    events = MemoryEventRepository([event])
    registrations = MemoryRegistrationRepository()
    workbooks = YandexDiskShowcaseRepository(disk_store)
    await workbooks.create_event_workbook(event.disk_resource_path)
    catalog = RegistrationCatalog(events, registrations, workbooks)
    administration = EventAdministrationService(events, workbooks, catalog, users, secret)
    export = await administration.create_pass_table(event.event_id)
    stored_event = await events.get(event.event_id)
    assert stored_event is not None and stored_event.pass_table_resource_path is not None
    pass_path = stored_event.pass_table_resource_path

    mutations = OrderedMutationService(events, catalog, users, secret)
    key = participant_key(secret, Platform.TELEGRAM, 11, event.event_id)

    def mutation(operation: Operation, payload: EnlistPayload | CharacterWishPayload | EmptyPayload):
        return OrderedRegistrationCommand(
            operation_id=str(uuid4()),
            event_id=event.event_id,
            operation=operation,
            platform=Platform.TELEGRAM,
            platform_user_id=11,
            participant_key=key,
            payload=payload,
        )

    await mutations.apply(
        mutation(Operation.ENLIST, EnlistPayload(display_name="Иван Иванов", wish_play="Без пожеланий"))
    )
    await mutations.apply(mutation(Operation.CONFIRM, CharacterWishPayload(character_wish="Без пожеланий")))

    confirmed = load_workbook(BytesIO(disk_store.files[pass_path]))
    try:
        assert confirmed.active.max_row == 2
        assert confirmed.active["A2"].value == "Иванов"
    finally:
        confirmed.close()
    assert disk_store.public_urls[pass_path] == export.public_url
    assert disk_store.replace_count[pass_path] == 1

    await mutations.apply(mutation(Operation.CANCEL, EmptyPayload()))

    cancelled = load_workbook(BytesIO(disk_store.files[pass_path]))
    try:
        assert cancelled.active.max_row == 1
    finally:
        cancelled.close()
    refreshed_event = await events.get(event.event_id)
    assert refreshed_event is not None
    assert refreshed_event.pass_table_public_url == export.public_url
    assert disk_store.public_urls[pass_path] == export.public_url
    assert disk_store.replace_count[pass_path] == 2


@pytest.mark.asyncio
async def test_archived_game_keeps_its_registration_and_pass_tables(disk_store: MemoryDiskStore, event) -> None:
    events = MemoryEventRepository([event])
    tables = MemoryRegistrationRepository()
    users = MemoryUserRepository()
    workbooks = YandexDiskShowcaseRepository(disk_store)
    await workbooks.create_event_workbook(event.disk_resource_path)
    catalog = RegistrationCatalog(events, tables, workbooks)
    service = EventAdministrationService(events, workbooks, catalog, users, "secret")
    await service.create_pass_table(event.event_id)
    stored_event = await events.get(event.event_id)
    assert stored_event is not None and stored_event.pass_table_resource_path is not None

    await catalog.archive(stored_event)

    archived_event = await events.get(event.event_id)
    assert event.disk_resource_path in disk_store.files
    assert stored_event.pass_table_resource_path in disk_store.files
    assert archived_event is not None and archived_event.status is EventStatus.CLOSED

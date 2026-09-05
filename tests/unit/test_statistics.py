from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from uuid import uuid4
from zipfile import ZipFile

import httpx
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill

from larp_bot.adapters.memory import MemoryDeferredTransport
from larp_bot.adapters.yandex_disk.repository import YandexDiskRestClient
from larp_bot.adapters.yandex_disk.statistics import BACKUP_RE, YandexDiskStatisticsRepository, table_filename
from larp_bot.adapters.yandex_disk.statistics_workbook import StatisticsWorkbook
from larp_bot.adapters.ymq.client import QueueEnvelope
from larp_bot.application.conversation import ADMIN, STATS_GAME, STATS_LINK, STATS_SEASON, STATS_SOURCE
from larp_bot.application.services import DomainError, OrderedMutationService
from larp_bot.application.statistics import StatisticsService
from larp_bot.application.worker import OrderedWorker
from larp_bot.domain.models import AttendanceStatus, Operation, OrderedRegistrationCommand, Registration
from tests.conftest import MemoryDiskStore
from tests.unit.test_conversation import engine_setup, inbound
from tests.unit.test_worker import FakeConsumer


def template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "2025-2026"
    sheet.append(
        ["Игрок", "Игр на текущий момент", "Игр на 01.09.2025", *[f"{n} игр" for n in range(5, 51, 5)], "Старая игра"]
    )
    sheet.append(["Иванов Пётр", "=C2+SUM(N2:N2)", 4, True, *([False] * 9), 1])
    sheet.append(["Петрова Анна", "=C3+SUM(N3:N3)", 0, *([False] * 10), "❌"])
    sheet.append(["Всего", '=COUNTIF(B2:B3,">0")'])
    sheet.append(["Участники сезона", "=SUMPRODUCT(--(B2:B3>C2:C3))"])
    sheet.append(["Новички", "=B4-C4"])
    sheet.append(["Активные игроки сезона", "=SUMPRODUCT(--(B2:B3>C2:C3+3))"])
    sheet["A2"].fill = PatternFill("solid", fgColor="D9D2E9")
    sheet.conditional_formatting.add(
        "N2:N3", CellIsRule(operator="equal", formula=["1"], fill=PatternFill("solid", fgColor="D9EAD3"))
    )
    sheet.conditional_formatting.add(
        "A2:A3", FormulaRule(formula=["AND($B2>=5,$D2=FALSE)"], fill=PatternFill("solid", fgColor="EAD1DC"))
    )
    historical = workbook.create_sheet("2024-2025")
    historical["A1"] = "Do not rewrite historical formula caches"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def sheet(content: bytes, *, formulas: bool = False):
    return load_workbook(BytesIO(content), data_only=not formulas).worksheets[0]


def test_rollover_retains_totals_awards_styles_and_historical_xml() -> None:
    original = template()
    result = StatisticsWorkbook(original).new_season(2026)
    assert result is not None
    current = sheet(result)
    assert current.title == "2026-2027"
    assert current.max_column == 13
    assert current["C1"].value == "Игр на 01.09.2026"
    assert current["B2"].value == current["C2"].value == 5
    assert current["D2"].value is True
    assert current["A2"].fill.fgColor.rgb == "00D9D2E9"
    assert sheet(result, formulas=True)["B2"].value == "=C2"
    assert StatisticsWorkbook(result).new_season(2026) is None
    with ZipFile(BytesIO(original)) as before, ZipFile(BytesIO(result)) as after:
        for path in before.namelist():
            if path.startswith("xl/worksheets/") or path == "xl/styles.xml":
                assert before.read(path) == after.read(path)


def test_game_matches_names_appends_before_summary_and_expands_formulas_and_colours() -> None:
    content = StatisticsWorkbook(template()).new_season(2026)
    assert content is not None
    result = StatisticsWorkbook(content).mark_game(
        "=Безопасное название", ["  ИВАНОВ  Петр ", "Новый Игрок", "Новый Игрок"]
    )
    assert result is not None
    current = sheet(result)
    assert current["N1"].value == "=Безопасное название"
    assert sheet(result, formulas=True)["N1"].data_type == "s"
    assert current["N2"].value == current["N4"].value == 1
    assert current["N3"].value is None
    assert current["B2"].value == 6
    assert current["A4"].value == "Новый Игрок"
    assert current["C4"].value == 0
    assert current["D4"].value is False
    assert current["A5"].value == "Всего"
    assert current["N5"].value == 2
    assert current["B6"].value == 2  # season participants
    assert current["B7"].value == 1  # newcomers
    assert any(str(cf.sqref) == "N2:N4" for cf in current.conditional_formatting)
    assert any(str(cf.sqref) == "A2:A4" for cf in current.conditional_formatting)
    assert StatisticsWorkbook(result).mark_game("=Безопасное название", ["Другой Игрок"]) is None
    for index in range(32):
        result = StatisticsWorkbook(result).mark_game(f"Игра {index}", ["Новый Игрок"])
        assert result is not None
    assert sheet(result)["B4"].value == 33  # beyond the original hardcoded AP range


@pytest.mark.parametrize("name", ["../initial", "disk:/x", "x/y", "showcase", "current", "pending", "bad\nfile"])
def test_source_name_cannot_escape_stats_or_select_internal_files(name: str) -> None:
    with pytest.raises(DomainError):
        table_filename(name)


async def test_backups_retention_source_selection_and_constant_url(disk_store: MemoryDiskStore) -> None:
    repo = YandexDiskStatisticsRepository(disk_store, retention=2)
    initial = template()
    disk_store.files[repo.path("initial.xlsx")] = initial
    url = await repo.link()
    await repo.select("initial", str(uuid4()))
    for index in range(4):
        before = disk_store.files[repo.path("current.xlsx")]
        op = str(uuid4())
        stamp = f"2026090500000000000{index}"
        assert await repo.edit(op, stamp, lambda book, i=index: book.mark_game(f"Игра {i}", ["Новый Игрок"]))
        assert disk_store.files[repo.path(f"backup-{stamp}-{op}.xlsx")] == before
        assert not await repo.edit(op, stamp, lambda book, i=index: book.mark_game(f"Игра {i}", ["Новый Игрок"]))
        assert await repo.link() == url
    backups = sorted(p for p in disk_store.files if BACKUP_RE.fullmatch(p.rsplit("/", 1)[1]))
    assert len(backups) == 2
    assert disk_store.files[repo.path("initial.xlsx")] == initial
    restore = disk_store.files[backups[0]]
    await repo.select(backups[0].rsplit("/", 1)[1], str(uuid4()))
    assert disk_store.files[repo.path("current.xlsx")] == restore
    assert disk_store.files[repo.path("showcase.xlsx")] == restore
    assert await repo.link() == url


class FailingStore(MemoryDiskStore):
    fail_name: str | None = None

    async def replace(self, path: str, content: bytes) -> None:
        if path.endswith("/" + str(self.fail_name)):
            self.fail_name = None
            raise RuntimeError("Disk temporarily unavailable")
        await super().replace(path, content)


@pytest.mark.parametrize("failure", ["current.xlsx", "showcase.xlsx", "state.json"])
async def test_interrupted_commit_recovers_without_double_counting(failure: str) -> None:
    store = FailingStore()
    repo = YandexDiskStatisticsRepository(store)
    store.files[repo.path("initial.xlsx")] = template()
    await repo.select("initial", str(uuid4()))
    url = await repo.link()
    store.fail_name = failure
    op, stamp = str(uuid4()), "20260905000000000000"
    with pytest.raises(RuntimeError):
        await repo.edit(op, stamp, lambda book: book.mark_game("Новая игра", ["Новый Игрок"]))
    # Simulate a fresh worker instance after a timeout/crash.
    repo = YandexDiskStatisticsRepository(store)
    await repo.edit(op, stamp, lambda book: book.mark_game("Новая игра", ["Новый Игрок"]))
    assert sheet(store.files[repo.path("current.xlsx")])["B4"].value == 1
    assert store.files[repo.path("current.xlsx")] == store.files[repo.path("showcase.xlsx")]
    assert len([p for p in store.files if "/backup-" in p]) == 1
    assert await repo.link() == url


async def test_admin_actions_use_global_fifo_and_confirmed_cyrillic_registration_names(disk_store, event) -> None:
    engine, _users, publisher, registrations = await engine_setup(disk_store, event, admin=True)
    menu = await engine.handle(inbound(1, ADMIN))
    assert {STATS_SOURCE, STATS_SEASON, STATS_LINK} <= {b.value for b in menu.buttons}
    game = await engine.handle(inbound(2, f"ag:manage:{event.event_id}"))
    assert STATS_GAME in {b.label for b in game.buttons}
    await engine.handle(inbound(3, STATS_SOURCE))
    await engine.handle(inbound(4, "initial"))
    await engine.handle(inbound(5, STATS_SEASON))
    await engine.handle(inbound(6, STATS_LINK))
    await engine.handle(inbound(7, f"ag:stats:{event.event_id}"))
    commands = publisher.commands
    assert len(commands) == 4
    assert all(c.event_id == "statistics" and c.operation is Operation.STATISTICS for c in commands)
    for status, name in (
        (AttendanceStatus.CONFIRMED, "Новый Игрок"),
        (AttendanceStatus.WAITING, "Ждущий Игрок"),
        (AttendanceStatus.CANCELLED, "Бывший Игрок"),
    ):
        await registrations.import_missing(
            [
                Registration(
                    event_id=event.event_id,
                    participant_key=uuid4().hex * 2,
                    display_name=name,
                    attendance_status=status,
                    wish_play="Без пожеланий",
                )
            ]
        )
    repo = YandexDiskStatisticsRepository(disk_store)
    disk_store.files[repo.path("initial.xlsx")] = template()
    service = StatisticsService(repo, engine.admins, engine.administration)
    command = OrderedRegistrationCommand.model_validate_json(commands[-1].model_dump_json())
    await service.apply(command)
    current = sheet(disk_store.files[repo.path("current.xlsx")])
    assert current["A4"].value == "Новый Игрок"
    assert current["A5"].value == "Всего"
    assert current["O4"].value == 1
    assert not await repo.edit(
        str(uuid4()),
        datetime.now(UTC).strftime("%Y%m%d%H%M%S%f"),
        lambda book: book.mark_game(event.name, ["Другой Игрок"]),
    )


async def test_stats_actions_denied_to_gamemasters_even_with_forged_callbacks(disk_store, event) -> None:
    engine, _users, publisher, _registrations = await engine_setup(disk_store, event, gamemaster=True)
    menu = await engine.handle(inbound(1, ADMIN))
    assert STATS_LINK not in {b.value for b in menu.buttons}
    for index, action in enumerate((STATS_SOURCE, STATS_SEASON, STATS_LINK, f"ag:stats:{event.event_id}"), start=2):
        response = await engine.handle(inbound(index, action))
        assert "прав" in response.text or "ведущ" in response.text
    assert not publisher.commands


def test_supplied_workbook_when_available() -> None:
    path = Path(__file__).resolve().parents[2] / "Копия Сводная статистика посещений.xlsx"
    if not path.exists():
        pytest.skip("Private sample workbook is not committed to the repository")
    original = path.read_bytes()
    result = StatisticsWorkbook(original).new_season(2026)
    assert result is not None
    assert sheet(result)["C4"].value == 4
    assert sheet(result)["A748"].value == "Дорогавцев Александр"
    result = StatisticsWorkbook(result).mark_game("Проверка", ["Новый Игрок", "Абдулазизова Софья"])
    assert result is not None
    assert sheet(result)["B4"].value == 5
    with ZipFile(BytesIO(original)) as before, ZipFile(BytesIO(result)) as after:
        for name in before.namelist():
            if name.startswith("xl/worksheets/"):
                assert before.read(name) == after.read(name)


def test_invalid_workbook_and_unrecognized_footer_are_rejected() -> None:
    with pytest.raises(DomainError):
        StatisticsWorkbook(b"not a workbook")
    book = load_workbook(BytesIO(template()))
    book.active["A10"] = "Additional data that must not be lost"
    output = BytesIO()
    book.save(output)
    with pytest.raises(DomainError):
        StatisticsWorkbook(output.getvalue())


async def test_missing_source_and_failed_backup_leave_existing_files_untouched() -> None:
    class BackupFailure(MemoryDiskStore):
        async def upload_new(self, path: str, content: bytes) -> None:
            if "/backup-" in path:
                raise RuntimeError("backup upload failed")
            await super().upload_new(path, content)

    store = BackupFailure()
    repo = YandexDiskStatisticsRepository(store)
    with pytest.raises(DomainError, match="не найден"):
        await repo.link()
    assert not store.files
    store.files[repo.path("initial.xlsx")] = template()
    await repo.select("initial", str(uuid4()))
    before = dict(store.files)
    with pytest.raises(RuntimeError, match="backup upload failed"):
        await repo.edit(str(uuid4()), "20260905000000000000", lambda book: book.new_season(2026))
    assert store.files == before


async def test_statistics_worker_delivers_link_without_event_lookup_and_rechecks_admin(disk_store, event) -> None:
    engine, users, publisher, _registrations = await engine_setup(disk_store, event, admin=True)
    await engine.handle(inbound(1, STATS_LINK))
    repo = YandexDiskStatisticsRepository(disk_store)
    disk_store.files[repo.path("initial.xlsx")] = template()
    service = StatisticsService(repo, engine.admins, engine.administration)
    command = publisher.commands[0]
    transport = MemoryDeferredTransport()
    consumer = FakeConsumer([QueueEnvelope(command, "receipt")])
    worker = OrderedWorker(
        consumer,
        OrderedMutationService(engine.events, engine.administration.catalog),
        users,
        transport,
        statistics=service,
    )
    assert await worker.run() == 1
    assert consumer.deleted == ["receipt"]
    assert "https://disk.example/" in transport.sent[0][3]
    before = dict(disk_store.files)
    unauthorized = command.model_copy(update={"platform_user_id": 999})
    with pytest.raises(DomainError, match="прав"):
        await service.apply(unauthorized)
    assert disk_store.files == before


async def test_disk_listing_paginates_and_exists_does_not_hide_api_errors() -> None:
    offsets = []

    def respond(request: httpx.Request) -> httpx.Response:
        if request.url.params.get("fields"):
            status = 404 if request.url.params["path"].endswith("missing") else 403
            return httpx.Response(status, request=request)
        offset = int(request.url.params["offset"])
        offsets.append(offset)
        items = [{"path": f"disk:/larp-bot/stats/backup-{offset}.xlsx", "type": "file"}]
        return httpx.Response(200, json={"_embedded": {"items": items, "total": 2}}, request=request)

    async with httpx.AsyncClient(transport=httpx.MockTransport(respond)) as client:
        disk = YandexDiskRestClient("test-token", client)
        assert len(await disk.list_files("disk:/larp-bot/stats")) == 2
        assert offsets == [0, 1]
        assert not await disk.exists("disk:/larp-bot/stats/missing")
        with pytest.raises(httpx.HTTPStatusError):
            await disk.exists("disk:/larp-bot/stats/forbidden")

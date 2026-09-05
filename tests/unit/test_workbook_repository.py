from __future__ import annotations

from datetime import UTC, datetime, timedelta
from io import BytesIO

import httpx
import pytest
from openpyxl import Workbook, load_workbook

from larp_bot.adapters.memory import MemoryEventRepository, MemoryRegistrationRepository
from larp_bot.adapters.yandex_disk.repository import (
    LEGACY_HEADERS,
    STATEFUL_HEADERS,
    VISIBLE_HEADERS,
    YandexDiskRestClient,
    YandexDiskShowcaseRepository,
)
from larp_bot.application.services import RegistrationCatalog
from larp_bot.domain.models import AttendanceStatus, Event, Registration
from tests.conftest import MemoryDiskStore


async def initialized(store: MemoryDiskStore, event: Event) -> YandexDiskShowcaseRepository:
    repository = YandexDiskShowcaseRepository(store)
    public_url = await repository.create_event_workbook(event.disk_resource_path)
    assert public_url == "https://disk.example/public/1"
    return repository


@pytest.mark.asyncio
async def test_new_workbook_is_a_visible_only_showcase(disk_store: MemoryDiskStore, event: Event) -> None:
    await initialized(disk_store, event)
    workbook = load_workbook(BytesIO(disk_store.files[event.disk_resource_path]))
    try:
        headers = tuple(cell.value for cell in workbook.active[1])
        assert headers == VISIBLE_HEADERS
        assert headers[2:4] == ("Профиль ВКонтакте", "Профиль в Telegram")
        assert workbook.active.max_column == 9
        assert not any(
            workbook.active.column_dimensions[column].hidden for column in ("A", "B", "C", "D", "E", "F", "G", "H", "I")
        )
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_showcase_is_regenerated_from_registration_rows(disk_store: MemoryDiskStore, event: Event) -> None:
    showcase = await initialized(disk_store, event)
    registration = Registration(
        event_id=event.event_id,
        participant_key="a" * 43,
        display_name='=HYPERLINK("bad")',
        vk_profile="https://vk.com/player",
        telegram_profile="https://t.me/player",
        wish_play="+SUM(1,2)",
        larp_experience=True,
        crossplay=False,
        character_wish="Doctor",
        attendance_status=AttendanceStatus.CONFIRMED,
    )

    await showcase.replace(event, [registration])

    workbook = load_workbook(BytesIO(disk_store.files[event.disk_resource_path]), data_only=False)
    try:
        assert tuple(workbook.active.cell(2, column).value for column in range(1, 10)) == (
            1,
            '\'=HYPERLINK("bad")',
            "https://vk.com/player",
            "https://t.me/player",
            "Да",
            "Нет",
            "'+SUM(1,2)",
            "Doctor",
            AttendanceStatus.CONFIRMED.value,
        )
        assert workbook.active.max_column == 9
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_cancelled_registrations_are_omitted_from_showcase(disk_store: MemoryDiskStore, event: Event) -> None:
    showcase = await initialized(disk_store, event)
    registrations = [
        Registration(
            event_id=event.event_id,
            participant_key="a" * 43,
            display_name="Cancelled player",
            wish_play="Anyone",
            attendance_status=AttendanceStatus.CANCELLED,
        ),
        Registration(
            event_id=event.event_id,
            participant_key="b" * 43,
            display_name="Active player",
            wish_play="Anyone",
        ),
    ]

    await showcase.replace(event, registrations)

    workbook = load_workbook(BytesIO(disk_store.files[event.disk_resource_path]))
    try:
        rows = [tuple(row) for row in workbook.active.iter_rows(min_row=2, max_col=2, values_only=True)]
        assert rows == [(1, "Active player")]
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_showcase_orders_rows_by_first_signup_even_when_input_is_arbitrary(
    disk_store: MemoryDiskStore, event: Event
) -> None:
    showcase = await initialized(disk_store, event)
    first_signup = datetime(2026, 9, 1, 10, tzinfo=UTC)
    registrations = [
        Registration(
            event_id=event.event_id,
            participant_key="c" * 43,
            display_name="Newest",
            wish_play="Anyone",
            created_at=first_signup + timedelta(hours=2),
        ),
        Registration(
            event_id=event.event_id,
            participant_key="a" * 43,
            display_name="Oldest, recently updated",
            wish_play="Anyone",
            created_at=first_signup,
            updated_at=first_signup + timedelta(days=1),
        ),
        Registration(
            event_id=event.event_id,
            participant_key="b" * 43,
            display_name="Middle",
            wish_play="Anyone",
            created_at=first_signup + timedelta(hours=1),
        ),
    ]

    await showcase.replace(event, registrations)

    workbook = load_workbook(BytesIO(disk_store.files[event.disk_resource_path]))
    try:
        rows = [tuple(row) for row in workbook.active.iter_rows(min_row=2, max_col=2, values_only=True)]
        assert rows == [
            (1, "Oldest, recently updated"),
            (2, "Middle"),
            (3, "Newest"),
        ]
    finally:
        workbook.close()


@pytest.mark.asyncio
@pytest.mark.parametrize("headers", [STATEFUL_HEADERS, LEGACY_HEADERS])
async def test_legacy_state_is_imported_once_then_removed_from_xlsx(
    headers: tuple[str, ...], disk_store: MemoryDiskStore, event: Event
) -> None:
    event.registrations_migrated_at = None
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    if headers == STATEFUL_HEADERS:
        sheet.append(
            (
                1,
                "Иван",
                "Да",
                "Нет",
                "Алиса",
                "Doctor",
                "Подтверждено",
                "a" * 43,
                "old-op",
                "2026-08-21T00:00:00+00:00",
            )
        )
    else:
        sheet.append(("Иван", "Алиса", "Doctor", "Подтверждено", "a" * 43, "old-op", "2026-08-21T00:00:00+00:00"))
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    disk_store.files[event.disk_resource_path] = output.getvalue()

    events = MemoryEventRepository([event])
    registrations = MemoryRegistrationRepository()
    showcase = YandexDiskShowcaseRepository(disk_store)
    catalog = RegistrationCatalog(events, registrations, showcase)
    migrated = await catalog.get(event, "a" * 43)

    assert migrated is not None
    assert migrated.character_wish == "Doctor"
    assert migrated.larp_experience is (True if headers == STATEFUL_HEADERS else None)
    stored_event = await events.get(event.event_id)
    assert stored_event is not None and stored_event.registrations_migrated_at is not None
    rendered = load_workbook(BytesIO(disk_store.files[event.disk_resource_path]))
    try:
        assert tuple(cell.value for cell in rendered.active[1]) == VISIBLE_HEADERS
        assert rendered.active.max_column == 9
    finally:
        rendered.close()


@pytest.mark.asyncio
async def test_optional_telegram_profile_is_rendered_as_blank(disk_store: MemoryDiskStore, event: Event) -> None:
    showcase = await initialized(disk_store, event)
    registration = Registration(
        event_id=event.event_id,
        participant_key="a" * 43,
        display_name="Player",
        vk_profile="https://vk.com/id1",
        telegram_profile=None,
        wish_play="Anyone",
    )

    await showcase.replace(event, [registration])

    workbook = load_workbook(BytesIO(disk_store.files[event.disk_resource_path]))
    try:
        assert workbook.active["C2"].value == "https://vk.com/id1"
        assert workbook.active["D2"].value is None
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_yandex_disk_download_follows_transfer_redirect() -> None:
    requests: list[str] = []

    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(str(request.url))
        if request.url.host == "cloud-api.yandex.net":
            return httpx.Response(200, json={"href": "https://downloader.disk.yandex.ru/file"})
        if request.url.host == "downloader.disk.yandex.ru":
            return httpx.Response(302, headers={"Location": "https://storage.yandex.net/file"})
        return httpx.Response(200, content=b"workbook")

    async with httpx.AsyncClient(transport=httpx.MockTransport(handler)) as client:
        disk = YandexDiskRestClient("token", client)
        assert await disk.download("disk:/larp-bot/events/event.xlsx") == b"workbook"

    assert [httpx.URL(url).host for url in requests] == [
        "cloud-api.yandex.net",
        "downloader.disk.yandex.ru",
        "storage.yandex.net",
    ]


@pytest.mark.asyncio
async def test_ydb_state_does_not_depend_on_showcase_contents(disk_store: MemoryDiskStore, event: Event) -> None:
    showcase = await initialized(disk_store, event)
    registration = Registration(
        event_id=event.event_id,
        participant_key="a" * 43,
        display_name="Player",
        wish_play="Anyone",
    )
    events = MemoryEventRepository([event])
    registrations = MemoryRegistrationRepository([registration])
    catalog = RegistrationCatalog(events, registrations, showcase)
    disk_store.files[event.disk_resource_path] = b"not an xlsx"

    found = await catalog.get(event, registration.participant_key)

    assert found == registration


def test_legacy_timestamp_is_timezone_aware() -> None:
    registration = Registration(
        event_id="event-a1",
        participant_key="a" * 43,
        display_name="Player",
        wish_play="Anyone",
        created_at=datetime(2026, 1, 1, tzinfo=UTC),
    )
    assert registration.created_at.tzinfo is UTC

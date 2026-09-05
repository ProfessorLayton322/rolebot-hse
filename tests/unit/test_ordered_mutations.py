from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from larp_bot.adapters.memory import MemoryEventRepository, MemoryRegistrationRepository, MemoryUserRepository
from larp_bot.adapters.yandex_disk.repository import PUBLIC_HEADERS, YandexDiskShowcaseRepository
from larp_bot.application.services import (
    OperationNotAllowed,
    OrderedMutationService,
    RegistrationCatalog,
)
from larp_bot.domain.models import (
    AttendanceStatus,
    BotIdentity,
    CharacterWishPayload,
    ConfirmationDeadlinePayload,
    EmptyPayload,
    EnlistPayload,
    Event,
    EventStatus,
    NotificationPayload,
    Operation,
    OrderedRegistrationCommand,
    Platform,
    TelegramUser,
)
from tests.conftest import MemoryDiskStore


def command(
    event: Event,
    operation: Operation,
    payload: EnlistPayload | CharacterWishPayload | ConfirmationDeadlinePayload | NotificationPayload | EmptyPayload,
    *,
    key: str = "a" * 43,
) -> OrderedRegistrationCommand:
    return OrderedRegistrationCommand(
        operation_id=str(uuid4()),
        event_id=event.event_id,
        operation=operation,
        platform=Platform.TELEGRAM,
        platform_user_id=1,
        participant_key=(
            None
            if operation
            in {
                Operation.OPEN_REGISTRATION,
                Operation.OPEN_CONFIRMATION,
                Operation.SEND_CONFIRMATION_REMINDER,
                Operation.SEND_CONFIRMED_NOTIFICATION,
                Operation.CLOSE_EVENT,
                Operation.DELETE_EVENT,
            }
            else key
        ),
        payload=payload,
    )


async def add_command_author_as_leader(events: MemoryEventRepository, event: Event) -> None:
    await events.add_leader(
        event.event_id,
        BotIdentity(platform=Platform.TELEGRAM, platform_user_id=1),
    )


@pytest.mark.asyncio
async def test_confirmed_notification_retry_appends_to_event_once(
    disk_store: MemoryDiskStore,
    event: Event,
) -> None:
    events = MemoryEventRepository([event])
    await add_command_author_as_leader(events, event)
    tables = MemoryRegistrationRepository()
    showcase = YandexDiskShowcaseRepository(disk_store)
    mutations = OrderedMutationService(events, RegistrationCatalog(events, tables, showcase))
    notification = command(
        event,
        Operation.SEND_CONFIRMED_NOTIFICATION,
        NotificationPayload(text="https://t.me/+GameChat_123"),
    )

    await mutations.apply(notification)
    await mutations.apply(notification)

    stored_event = await events.get(event.event_id)
    assert stored_event is not None
    assert stored_event.confirmed_notifications == ["https://t.me/+GameChat_123"]


@pytest.mark.asyncio
async def test_worker_rejects_privileged_mutation_from_nonleader(
    disk_store: MemoryDiskStore,
    event: Event,
) -> None:
    events = MemoryEventRepository([event])
    tables = MemoryRegistrationRepository()
    mutations = OrderedMutationService(
        events,
        RegistrationCatalog(events, tables, YandexDiskShowcaseRepository(disk_store)),
    )

    with pytest.raises(OperationNotAllowed, match="Только ведущие"):
        await mutations.apply(
            command(
                event,
                Operation.SEND_CONFIRMED_NOTIFICATION,
                NotificationPayload(text="Несанкционированное уведомление"),
            )
        )

    stored_event = await events.get(event.event_id)
    assert stored_event is not None and stored_event.confirmed_notifications == []


@pytest.mark.asyncio
async def test_legacy_enlist_command_backfills_profile_columns(disk_store: MemoryDiskStore, event: Event) -> None:
    tables = MemoryRegistrationRepository()
    showcase = YandexDiskShowcaseRepository(disk_store)
    await showcase.create_event_workbook(event.disk_resource_path)
    events = MemoryEventRepository([event])
    users = MemoryUserRepository()
    await users.save(
        TelegramUser(
            tg_id=1,
            full_name="Player",
            vk_url="https://vk.com/player",
            crossplay=False,
            larp_experience=True,
            needs_pass=False,
        )
    )
    mutations = OrderedMutationService(events, RegistrationCatalog(events, tables, showcase), users)

    await mutations.apply(command(event, Operation.ENLIST, EnlistPayload(display_name="Player", wish_play="A")))

    workbook = load_workbook(BytesIO(disk_store.files[event.disk_resource_path]))
    try:
        assert workbook.active["C2"].value == "https://vk.com/player"
        assert workbook.active["D2"].value is None
        assert workbook.active["E2"].value == "Да"
        assert workbook.active["F2"].value == "Нет"
    finally:
        workbook.close()
    stored_event = await events.get(event.event_id)
    assert stored_event is not None and stored_event.public_table_resource_path is not None
    public_workbook = load_workbook(BytesIO(disk_store.files[stored_event.public_table_resource_path]))
    try:
        assert tuple(cell.value for cell in public_workbook.active[1]) == PUBLIC_HEADERS
        assert public_workbook.active.max_column == 7
        assert "https://vk.com/player" not in {cell.value for row in public_workbook.active.iter_rows() for cell in row}
    finally:
        public_workbook.close()


@pytest.mark.asyncio
async def test_worker_rejects_enlist_when_ydb_profile_is_missing(disk_store: MemoryDiskStore, event: Event) -> None:
    tables = MemoryRegistrationRepository()
    showcase = YandexDiskShowcaseRepository(disk_store)
    await showcase.create_event_workbook(event.disk_resource_path)
    events = MemoryEventRepository([event])
    users = MemoryUserRepository()
    mutations = OrderedMutationService(events, RegistrationCatalog(events, tables, showcase), users)

    with pytest.raises(OperationNotAllowed, match="полностью заполните профиль"):
        await mutations.apply(command(event, Operation.ENLIST, EnlistPayload(display_name="Player", wish_play="A")))

    assert await tables.get(event.event_id, "a" * 43) is None


@pytest.mark.asyncio
async def test_enlist_persists_both_profiles_in_showcase(disk_store: MemoryDiskStore, event: Event) -> None:
    tables = MemoryRegistrationRepository()
    showcase = YandexDiskShowcaseRepository(disk_store)
    await showcase.create_event_workbook(event.disk_resource_path)
    events = MemoryEventRepository([event])
    mutations = OrderedMutationService(events, RegistrationCatalog(events, tables, showcase))
    payload = EnlistPayload(
        display_name="Player",
        wish_play="A",
        vk_profile="https://vk.com/player",
        telegram_profile="https://t.me/player",
    )

    await mutations.apply(command(event, Operation.ENLIST, payload))

    registration = await tables.get(event.event_id, "a" * 43)
    assert registration is not None
    assert registration.vk_profile == "https://vk.com/player"
    assert registration.telegram_profile == "https://t.me/player"
    workbook = load_workbook(BytesIO(disk_store.files[event.disk_resource_path]))
    try:
        assert workbook.active["C2"].value == "https://vk.com/player"
        assert workbook.active["D2"].value == "https://t.me/player"
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_fifo_sequence_has_required_final_state(disk_store: MemoryDiskStore, event: Event) -> None:
    tables = MemoryRegistrationRepository()
    showcase = YandexDiskShowcaseRepository(disk_store)
    await showcase.create_event_workbook(event.disk_resource_path)
    events = MemoryEventRepository([event])
    mutations = OrderedMutationService(events, RegistrationCatalog(events, tables, showcase))
    sequence = [
        command(event, Operation.ENLIST, EnlistPayload(display_name="Player", wish_play="A")),
        command(event, Operation.CONFIRM, CharacterWishPayload(character_wish="A")),
        command(event, Operation.UPDATE_CHARACTER_WISH, CharacterWishPayload(character_wish="B")),
        command(event, Operation.CANCEL, EmptyPayload()),
    ]
    for item in sequence:
        await mutations.apply(item)
    registration = await tables.get(event.event_id, "a" * 43)
    assert registration is not None
    assert registration.character_wish == "B"
    assert registration.attendance_status is AttendanceStatus.CANCELLED
    workbook = load_workbook(BytesIO(disk_store.files[event.disk_resource_path]))
    try:
        assert workbook.active.max_row == 1
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_fifo_character_update_preserves_confirmed_state(disk_store: MemoryDiskStore, event: Event) -> None:
    tables = MemoryRegistrationRepository()
    showcase = YandexDiskShowcaseRepository(disk_store)
    await showcase.create_event_workbook(event.disk_resource_path)
    events = MemoryEventRepository([event])
    mutations = OrderedMutationService(events, RegistrationCatalog(events, tables, showcase))
    for item in (
        command(event, Operation.ENLIST, EnlistPayload(display_name="Player", wish_play="A")),
        command(event, Operation.CONFIRM, CharacterWishPayload(character_wish="A")),
        command(event, Operation.UPDATE_CHARACTER_WISH, CharacterWishPayload(character_wish="B")),
    ):
        await mutations.apply(item)
    registration = await tables.get(event.event_id, "a" * 43)
    assert registration is not None
    assert registration.character_wish == "B"
    assert registration.attendance_status is AttendanceStatus.CONFIRMED


@pytest.mark.asyncio
async def test_created_game_allows_enlist_but_rejects_confirmation_until_admin_opens_it(
    disk_store: MemoryDiskStore, event: Event
) -> None:
    event.status = EventStatus.CREATED
    tables = MemoryRegistrationRepository()
    showcase = YandexDiskShowcaseRepository(disk_store)
    await showcase.create_event_workbook(event.disk_resource_path)
    events = MemoryEventRepository([event])
    mutations = OrderedMutationService(events, RegistrationCatalog(events, tables, showcase))
    await mutations.apply(command(event, Operation.ENLIST, EnlistPayload(display_name="User A", wish_play="X")))
    with pytest.raises(OperationNotAllowed, match="ещё не открыто"):
        await mutations.apply(command(event, Operation.CONFIRM, CharacterWishPayload(character_wish="Doctor")))

    deadline = datetime(2026, 9, 10, 16, tzinfo=UTC)
    await add_command_author_as_leader(events, event)
    await mutations.apply(command(event, Operation.OPEN_CONFIRMATION, ConfirmationDeadlinePayload(deadline=deadline)))
    await mutations.apply(command(event, Operation.CONFIRM, CharacterWishPayload(character_wish="Doctor")))
    registration = await tables.get(event.event_id, "a" * 43)
    assert registration is not None
    assert registration.attendance_status is AttendanceStatus.CONFIRMED


@pytest.mark.asyncio
async def test_close_orders_both_enlist_and_confirmation_rejection(disk_store: MemoryDiskStore, event: Event) -> None:
    tables = MemoryRegistrationRepository()
    showcase = YandexDiskShowcaseRepository(disk_store)
    await showcase.create_event_workbook(event.disk_resource_path)
    events = MemoryEventRepository([event])
    await add_command_author_as_leader(events, event)
    mutations = OrderedMutationService(events, RegistrationCatalog(events, tables, showcase))
    await mutations.apply(command(event, Operation.ENLIST, EnlistPayload(display_name="User A", wish_play="X")))
    await mutations.apply(command(event, Operation.CONFIRM, CharacterWishPayload(character_wish="Doctor")))
    await mutations.apply(command(event, Operation.CLOSE_EVENT, EmptyPayload()))
    with pytest.raises(OperationNotAllowed):
        await mutations.apply(
            command(
                event,
                Operation.ENLIST,
                EnlistPayload(display_name="User B", wish_play="X"),
                key="b" * 43,
            )
        )
    with pytest.raises(OperationNotAllowed, match="закрыто"):
        await mutations.apply(command(event, Operation.CONFIRM, CharacterWishPayload(character_wish="Changed")))
    registration = await tables.get(event.event_id, "a" * 43)
    assert registration is not None
    assert registration.attendance_status is AttendanceStatus.CONFIRMED
    await mutations.apply(command(event, Operation.UPDATE_CHARACTER_WISH, CharacterWishPayload(character_wish="Medic")))
    updated = await tables.get(event.event_id, "a" * 43)
    assert updated is not None
    assert updated.character_wish == "Medic"
    assert updated.attendance_status is AttendanceStatus.CONFIRMED


@pytest.mark.asyncio
async def test_admin_status_commands_allow_any_transition(disk_store: MemoryDiskStore, event: Event) -> None:
    event.status = EventStatus.CREATED
    tables = MemoryRegistrationRepository()
    showcase = YandexDiskShowcaseRepository(disk_store)
    await showcase.create_event_workbook(event.disk_resource_path)
    events = MemoryEventRepository([event])
    await add_command_author_as_leader(events, event)
    mutations = OrderedMutationService(events, RegistrationCatalog(events, tables, showcase))

    await mutations.apply(command(event, Operation.CLOSE_EVENT, EmptyPayload()))
    closed = await events.get(event.event_id)
    assert closed is not None and closed.status is EventStatus.CLOSED

    deadline = datetime(2026, 9, 10, 16, tzinfo=UTC)
    await mutations.apply(command(event, Operation.OPEN_CONFIRMATION, ConfirmationDeadlinePayload(deadline=deadline)))
    opened = await events.get(event.event_id)
    assert opened is not None and opened.status is EventStatus.CONFIRMATION_OPEN
    assert opened.confirmation_deadline == deadline
    await mutations.apply(command(event, Operation.ENLIST, EnlistPayload(display_name="User A", wish_play="X")))
    await mutations.apply(command(event, Operation.CONFIRM, CharacterWishPayload(character_wish="Doctor")))

    await mutations.apply(command(event, Operation.OPEN_REGISTRATION, EmptyPayload()))
    registration = await events.get(event.event_id)
    assert registration is not None and registration.status is EventStatus.CREATED
    with pytest.raises(OperationNotAllowed, match="ещё не открыто"):
        await mutations.apply(command(event, Operation.CONFIRM, CharacterWishPayload(character_wish="Changed")))


@pytest.mark.asyncio
async def test_legacy_delete_command_archives_without_deleting_game_data(
    disk_store: MemoryDiskStore, event: Event
) -> None:
    tables = MemoryRegistrationRepository()
    showcase = YandexDiskShowcaseRepository(disk_store)
    await showcase.create_event_workbook(event.disk_resource_path)
    events = MemoryEventRepository([event])
    await add_command_author_as_leader(events, event)
    mutations = OrderedMutationService(events, RegistrationCatalog(events, tables, showcase))
    await mutations.apply(command(event, Operation.ENLIST, EnlistPayload(display_name="User A", wish_play="X")))
    await mutations.apply(command(event, Operation.CONFIRM, CharacterWishPayload(character_wish="Doctor")))
    result = await mutations.apply(command(event, Operation.DELETE_EVENT, EmptyPayload()))

    stored_event = await events.get(event.event_id)
    stored_registration = await tables.get(event.event_id, "a" * 43)
    assert result == "Игра архивирована; таблицы и записи сохранены"
    assert stored_event is not None and stored_event.status is EventStatus.CLOSED
    assert stored_registration is not None and stored_registration.character_wish == "Doctor"
    assert event.disk_resource_path in disk_store.files

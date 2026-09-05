from __future__ import annotations

import asyncio
import json
from collections.abc import Sequence
from datetime import UTC, datetime
from io import BytesIO
from typing import Protocol

import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from larp_bot.domain.models import AttendanceStatus, Event, PassDetails, Registration

VISIBLE_HEADERS = (
    "№",
    "Имя",
    "Профиль ВКонтакте",
    "Профиль в Telegram",
    "Предыдущий опыт в LARP",
    "Готовность к кроссполу",
    "С кем хочу играть",
    "Пожелания по персонажу",
    "Текущий статус",
)
# These layouts are read only while migrating deployments that used XLSX as
# storage. Newly generated showcase files contain VISIBLE_HEADERS exclusively.
STATEFUL_HEADERS = (
    "№",
    "Имя",
    "Предыдущий опыт в LARP",
    "Готовность к кроссполу",
    "С кем хочу играть",
    "Пожелания по персонажу",
    "Текущий статус",
    "participant_key",
    "last_operation_id",
    "updated_at",
)
LEGACY_HEADERS = (
    "Имя",
    "С кем хочу играть",
    "Пожелания по персонажу",
    "Статус",
    "participant_key",
    "last_operation_id",
    "updated_at",
)
PASS_TABLE_HEADERS = (
    "Фамилия (Кириллицей)",
    "Имя (Кириллицей)",
    "Отчество (Кириллицей)",
    "Иностранный гражданин (Да/Нет)",
    "Фамилия (Латиницей)",
    "Имя (Латиницей)",
    "Отчество (Латиницей)",
    "Телефон",
    "E-mail",
)


class WorkbookIntegrityError(RuntimeError):
    pass


class DiskObjectStore(Protocol):
    async def download(self, path: str) -> bytes: ...

    async def upload_new(self, path: str, content: bytes) -> None: ...

    async def replace(self, path: str, content: bytes) -> None: ...

    async def publish(self, path: str) -> str: ...

    async def delete(self, path: str) -> None: ...


def _cell_text(value: object) -> str:
    return "" if value is None else str(value)


def safe_cell(value: str) -> str:
    """Neutralize spreadsheet formulas while retaining the visible user text."""
    return f"'{value}" if value.startswith(("=", "+", "-", "@")) else value


def display_cell(value: object) -> str:
    text = _cell_text(value)
    if len(text) > 1 and text[0] == "'" and text[1] in "=+-@":
        return text[1:]
    return text


def _bool_cell(value: bool | None) -> str:
    if value is None:
        return "Не указано"
    return "Да" if value else "Нет"


def _parse_bool(value: object) -> bool | None:
    text = _cell_text(value)
    if text == "Да":
        return True
    if text == "Нет":
        return False
    return None


def _format_sheet(sheet: Worksheet) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:I1"
    widths = (8, 30, 32, 32, 25, 26, 35, 45, 18)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width


def showcase_workbook_bytes(registrations: Sequence[Registration] = ()) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Регистрация"
    sheet.append(VISIBLE_HEADERS)
    # A repository is not required to return rows in insertion order. Keep the
    # public signup sheet chronological at the projection boundary so later
    # profile, confirmation, and cancellation updates cannot move a player.
    active_registrations = (
        registration
        for registration in registrations
        if registration.attendance_status is not AttendanceStatus.CANCELLED
    )
    ordered_registrations = sorted(active_registrations, key=lambda row: (row.created_at, row.participant_key))
    for number, registration in enumerate(ordered_registrations, start=1):
        sheet.append(
            (
                number,
                safe_cell(registration.display_name),
                safe_cell(registration.vk_profile),
                safe_cell(registration.telegram_profile or ""),
                _bool_cell(registration.larp_experience),
                _bool_cell(registration.crossplay),
                safe_cell(registration.wish_play),
                safe_cell(registration.character_wish),
                registration.attendance_status.value,
            )
        )
    _format_sheet(sheet)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def pass_table_workbook_bytes(profiles: Sequence[PassDetails] = ()) -> bytes:
    """Render the pass-list schema and header style used by the supplied venue template."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лист1"
    sheet.append(PASS_TABLE_HEADERS)
    thin = Side(style="thin", color="000000")
    header_fill = PatternFill(fill_type="solid", fgColor=Color(theme=5))
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(name="Calibri Light", size=14)
        cell.alignment = Alignment(wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell.number_format = "@"
    sheet.row_dimensions[1].height = 57
    widths = (24.85546875, 22.140625, 24.28515625, 13, 40.28515625, 22.140625, 13, 33.85546875, 43.7109375)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width

    for profile in profiles:
        sheet.append(
            (
                profile.surname_cyrillic,
                profile.name_cyrillic,
                profile.patronym_cyrillic,
                "Да" if profile.foreigner else "Нет",
                profile.surname_latin if profile.foreigner else None,
                profile.name_latin if profile.foreigner else None,
                profile.patronym_latin if profile.foreigner else None,
                profile.mobile_phone,
                profile.email,
            )
        )
        sheet.row_dimensions[sheet.max_row].height = 19.5
        for cell in sheet[sheet.max_row]:
            cell.font = Font(name="Calibri Light", size=14)
            cell.alignment = Alignment(wrap_text=True)
            cell.number_format = "@"

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _legacy_registrations(event_id: str, content: bytes) -> list[Registration]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    except Exception as exc:
        raise WorkbookIntegrityError("registration resource is not a valid XLSX") from exc
    try:
        sheet = workbook.active
        actual = tuple(_cell_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1)))
        if actual not in {STATEFUL_HEADERS, LEGACY_HEADERS}:
            raise WorkbookIntegrityError(
                f"unexpected legacy registration workbook schema: {json.dumps(actual, ensure_ascii=False)}"
            )
        registrations: list[Registration] = []
        stateful = actual == STATEFUL_HEADERS
        for row in sheet.iter_rows(min_row=2, values_only=True):
            key_index = 7 if stateful else 4
            participant = _cell_text(row[key_index] if len(row) > key_index else None)
            if not participant:
                continue
            updated_index = 9 if stateful else 6
            updated_raw = _cell_text(row[updated_index] if len(row) > updated_index else None)
            try:
                updated_at = datetime.fromisoformat(updated_raw)
            except ValueError:
                updated_at = datetime.now(UTC)
            registrations.append(
                Registration(
                    event_id=event_id,
                    participant_key=participant,
                    display_name=display_cell(row[1] if stateful else row[0]),
                    larp_experience=_parse_bool(row[2]) if stateful else None,
                    crossplay=_parse_bool(row[3]) if stateful else None,
                    wish_play=display_cell(row[4] if stateful else row[1]),
                    character_wish=display_cell(row[5] if stateful else row[2]),
                    attendance_status=AttendanceStatus(_cell_text(row[6] if stateful else row[3])),
                    last_operation_id=_cell_text(row[8] if stateful else row[5]),
                    created_at=updated_at,
                    updated_at=updated_at,
                )
            )
        return registrations
    finally:
        workbook.close()


class YandexDiskRestClient:
    API = "https://cloud-api.yandex.net/v1/disk"

    def __init__(self, token: str, client: httpx.AsyncClient | None = None) -> None:
        self._client = client or httpx.AsyncClient(timeout=httpx.Timeout(30.0))
        self._headers = {"Authorization": f"OAuth {token}"}

    async def _link(self, endpoint: str, path: str, *, overwrite: bool = False) -> str:
        response = await self._client.get(
            f"{self.API}/{endpoint}",
            headers=self._headers,
            params={"path": path, "overwrite": str(overwrite).lower()},
        )
        response.raise_for_status()
        href = response.json().get("href")
        if not isinstance(href, str) or not href.startswith("https://"):
            raise RuntimeError("Yandex Disk did not return an HTTPS transfer URL")
        return href

    async def download(self, path: str) -> bytes:
        href = await self._link("resources/download", path)
        response = await self._client.get(href, follow_redirects=True)
        response.raise_for_status()
        return response.content

    async def _upload(self, path: str, content: bytes, *, overwrite: bool) -> None:
        href = await self._link("resources/upload", path, overwrite=overwrite)
        response = await self._client.put(
            href,
            content=content,
            headers={"Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        )
        response.raise_for_status()

    async def _ensure_parent_directories(self, path: str) -> None:
        if not path.startswith("disk:/") or "/" not in path.removeprefix("disk:/"):
            return
        parent = path.rsplit("/", 1)[0]
        segments = parent.removeprefix("disk:/").split("/")
        for index in range(1, len(segments) + 1):
            directory = "disk:/" + "/".join(segments[:index])
            response = await self._client.put(
                f"{self.API}/resources",
                headers=self._headers,
                params={"path": directory},
            )
            if response.status_code != 409:
                response.raise_for_status()

    async def upload_new(self, path: str, content: bytes) -> None:
        await self._ensure_parent_directories(path)
        await self._upload(path, content, overwrite=False)

    async def replace(self, path: str, content: bytes) -> None:
        await self._upload(path, content, overwrite=True)

    async def publish(self, path: str) -> str:
        response = await self._client.put(f"{self.API}/resources/publish", headers=self._headers, params={"path": path})
        response.raise_for_status()
        metadata = await self._client.get(
            f"{self.API}/resources", headers=self._headers, params={"path": path, "fields": "public_url"}
        )
        metadata.raise_for_status()
        public_url = metadata.json().get("public_url")
        if not isinstance(public_url, str) or not public_url.startswith("https://"):
            raise RuntimeError("Yandex Disk resource was published without a public URL")
        return public_url

    async def delete(self, path: str) -> None:
        response = await self._client.delete(
            f"{self.API}/resources",
            headers=self._headers,
            params={"path": path, "permanently": "true"},
        )
        if response.status_code != 404:
            response.raise_for_status()


class YandexDiskShowcaseRepository:
    def __init__(self, store: DiskObjectStore) -> None:
        self.store = store

    async def create_event_workbook(self, disk_path: str) -> str:
        await self.store.upload_new(disk_path, showcase_workbook_bytes())
        try:
            return await self.store.publish(disk_path)
        except Exception:
            await self.store.delete(disk_path)
            raise

    async def delete_event_workbook(self, disk_path: str) -> None:
        await self.store.delete(disk_path)

    async def create_pass_table(self, disk_path: str, profiles: Sequence[PassDetails]) -> str:
        content = await asyncio.to_thread(pass_table_workbook_bytes, profiles)
        await self.store.upload_new(disk_path, content)
        try:
            return await self.store.publish(disk_path)
        except Exception:
            await self.store.delete(disk_path)
            raise

    async def delete_pass_table(self, disk_path: str) -> None:
        await self.store.delete(disk_path)

    async def read_legacy_registrations(self, event: Event) -> Sequence[Registration]:
        content = await self.store.download(event.disk_resource_path)
        return await asyncio.to_thread(_legacy_registrations, event.event_id, content)

    async def replace(self, event: Event, registrations: Sequence[Registration]) -> None:
        content = await asyncio.to_thread(showcase_workbook_bytes, registrations)
        await self.store.replace(event.disk_resource_path, content)

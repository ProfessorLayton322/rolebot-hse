"""Edit only the current season's OOXML, retaining historical sheets and caches.

The supplied Google Sheets export contains cached DUMMYFUNCTION/array formulas.
An openpyxl round trip discards those caches in *every* season. Keeping the ZIP
parts intact also preserves styles, checkbox metadata and unrelated worksheets.
"""

from __future__ import annotations

import re
from collections.abc import Sequence
from copy import deepcopy
from io import BytesIO
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, BadZipFile, ZipFile

from openpyxl.utils import column_index_from_string, get_column_letter

from larp_bot.application.services import DomainError

NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG = "http://schemas.openxmlformats.org/package/2006/relationships"
CT = "http://schemas.openxmlformats.org/package/2006/content-types"
ET.register_namespace("", NS)
ET.register_namespace("r", REL)


def tag(name: str) -> str:
    return f"{{{NS}}}{name}"


def normalized_name(value: str) -> str:
    return " ".join(value.split()).casefold().replace("ё", "е")


class StatisticsWorkbook:
    def __init__(self, content: bytes) -> None:
        self.original = content
        try:
            with ZipFile(BytesIO(content)) as archive:
                self.parts = {name: archive.read(name) for name in archive.namelist()}
            self.book = ET.fromstring(self.parts["xl/workbook.xml"])  # noqa: S314
            self.rels = ET.fromstring(self.parts["xl/_rels/workbook.xml.rels"])  # noqa: S314
            sheets = self.book.find(tag("sheets"))
            assert sheets is not None
            self.sheets = sheets
            seasons = [s for s in self.sheets if re.fullmatch(r"20\d\d-20\d\d", s.attrib["name"])]
            self.season = max(seasons, key=lambda s: s.attrib["name"])
            target = next(r.attrib["Target"] for r in self.rels if r.attrib["Id"] == self.season.attrib[f"{{{REL}}}id"])
            self.path = target.lstrip("/") if target.startswith("/") else f"xl/{target}"
            self.sheet = ET.fromstring(self.parts[self.path])  # noqa: S314
            data = self.sheet.find(tag("sheetData"))
            assert data is not None
            self.data = data
            self.strings: list[str] = []
            if "xl/sharedStrings.xml" in self.parts:
                strings = ET.fromstring(self.parts["xl/sharedStrings.xml"])  # noqa: S314
                self.strings = ["".join(t.text or "" for t in s.iter(tag("t"))) for s in strings]
            self.rows = {int(r.attrib["r"]): r for r in self.data}
            self.cells = {c.attrib["r"]: c for r in self.data for c in r}
            if self.value("A1") != "Игрок" or self.value("D1") != "5 игр" or self.value("M1") != "50 игр":
                raise ValueError("unsupported layout")
            self.last_column = max(
                column_index_from_string(re.sub(r"\d", "", c.attrib["r"]))
                for c in self.rows[1]
                if self.value(c.attrib["r"]) not in (None, "")
            )
            self.last_player = 1
            for row in sorted(self.rows):
                if row == 1:
                    continue
                name = self.value(f"A{row}")
                cell = self.cells.get(f"A{row}")
                if not isinstance(name, str) or name in {
                    "Всего",
                    "Участники сезона",
                    "Новички",
                    "Активные игроки сезона",
                }:
                    break
                if cell is not None and cell.find(tag("f")) is not None:
                    break
                self.last_player = row
            footer_labels = {"Всего", "Участники сезона", "Новички", "Активные игроки сезона"}
            for index, xml_row in self.rows.items():
                if index <= self.last_player:
                    continue
                populated = any(
                    self.value(c.attrib["r"]) not in (None, "") or c.find(tag("f")) is not None for c in xml_row
                )
                if not populated:
                    continue
                name = self.value(f"A{index}")
                cell = self.cells.get(f"A{index}")
                exported_total = index == self.last_player + 1 and cell is not None and cell.find(tag("f")) is not None
                if index > self.last_player + 4 or (name not in footer_labels and not exported_total):
                    raise ValueError("unexpected data below player rows")
        except (BadZipFile, KeyError, ValueError, AssertionError, StopIteration, ET.ParseError) as exc:
            raise DomainError("Нужна XLSX-таблица с сезонным листом, колонками Игрок, итоги и 5–50 игр.") from exc

    def value(self, address: str) -> str | float | bool | None:
        cell = self.cells.get(address)
        if cell is None:
            return None
        kind = cell.get("t")
        value = cell.findtext(tag("v"))
        if kind == "inlineStr":
            return "".join(t.text or "" for t in cell.iter(tag("t")))
        if value is None:
            return None
        if kind == "s":
            return self.strings[int(value)]
        if kind in {"str", "e"}:
            return value
        if kind == "b":
            return value == "1"
        return float(value)

    def number(self, address: str, *, sum_cell: bool = False) -> float:
        value = self.value(address)
        cell = self.cells.get(address)
        if isinstance(value, (float, bool)):
            return float(value)
        # SUM(range) ignores text markers such as the supplied sheet's ❌.
        if sum_cell and isinstance(value, str) and (cell is None or cell.get("t") != "e"):
            return 0
        if value in (None, "") and (cell is None or cell.find(tag("f")) is None):
            return 0
        raise DomainError(f"Нет числового результата в {self.season.attrib['name']}!{address}. Пересчитайте XLSX.")

    def total(self, row: int) -> float:
        return self.number(f"C{row}") + sum(
            self.number(f"{get_column_letter(c)}{row}", sum_cell=True) for c in range(14, self.last_column + 1)
        )

    def set_cell(
        self, row: int, col: int, value: str | float | bool, *, style: str | None = None, formula: str | None = None
    ) -> None:
        address = f"{get_column_letter(col)}{row}"
        if row not in self.rows:
            self.rows[row] = ET.SubElement(self.data, tag("row"), r=str(row))
        cell = self.cells.get(address)
        if cell is None:
            cell = ET.SubElement(self.rows[row], tag("c"), r=address)
            self.cells[address] = cell
        if style is not None:
            cell.set("s", style)
        for child in list(cell):
            cell.remove(child)
        cell.attrib.pop("t", None)
        if formula:
            ET.SubElement(cell, tag("f")).text = formula
        if isinstance(value, str):
            cell.set("t", "inlineStr")
            ET.SubElement(ET.SubElement(cell, tag("is")), tag("t")).text = value
        else:
            if isinstance(value, bool):
                cell.set("t", "b")
            ET.SubElement(cell, tag("v")).text = str(int(value) if isinstance(value, bool) else value)

    def style(self, row: int, col: int) -> str:
        cell = self.cells.get(f"{get_column_letter(col)}{row}")
        return cell.get("s", "0") if cell is not None else "0"

    def has_game(self, name: str) -> bool:
        return any(
            normalized_name(str(self.value(f"{get_column_letter(c)}1") or "")) == normalized_name(name)
            for c in range(14, self.last_column + 1)
        )

    def new_season(self, year: int) -> bytes | None:
        title = f"{year}-{year + 1}"
        if any(s.attrib["name"] == title for s in self.sheets):
            return None
        if title < self.season.attrib["name"]:
            raise DomainError("В таблице уже есть более поздний сезон.")
        totals = {r: self.total(r) for r in range(2, self.last_player + 1)}
        # Start with the fixed columns only, carrying names, totals and award flags.
        for xml_row in list(self.data):
            index = int(xml_row.attrib["r"])
            if index > self.last_player:
                self.data.remove(xml_row)
                self.rows.pop(index, None)
            else:
                for cell in list(xml_row):
                    if column_index_from_string(re.sub(r"\d", "", cell.attrib["r"])) > 13:
                        xml_row.remove(cell)
        self.cells = {c.attrib["r"]: c for r in self.data for c in r}
        self.set_cell(1, 3, f"Игр на 01.09.{year}")
        for row, total in totals.items():
            self.set_cell(row, 3, total)
        self.last_column = 13
        # A new sheet must not inherit relationships to old comments/drawings.
        for node in list(self.sheet):
            if node.tag not in {
                tag(n)
                for n in (
                    "sheetPr",
                    "dimension",
                    "sheetViews",
                    "sheetFormatPr",
                    "cols",
                    "sheetData",
                    "conditionalFormatting",
                    "pageMargins",
                    "pageSetup",
                    "printOptions",
                )
            }:
                self.sheet.remove(node)
        columns = self.sheet.find(tag("cols"))
        if columns is not None:
            for col in list(columns):
                if int(col.attrib["min"]) > 13:
                    columns.remove(col)
                elif int(col.attrib["max"]) > 13:
                    col.set("max", "13")
        index = max(int(s.attrib["sheetId"]) for s in self.sheets) + 1
        self.path = f"xl/worksheets/statsSeason{index}.xml"
        rid = f"statsSeason{index}"
        ET.SubElement(
            self.rels,
            f"{{{PKG}}}Relationship",
            Id=rid,
            Type=f"{REL}/worksheet",
            Target=f"worksheets/statsSeason{index}.xml",
        )
        self.sheets.insert(0, ET.Element(tag("sheet"), name=title, sheetId=str(index), attrib={f"{{{REL}}}id": rid}))
        types = ET.fromstring(self.parts["[Content_Types].xml"])  # noqa: S314
        ET.SubElement(
            types,
            f"{{{CT}}}Override",
            PartName=f"/{self.path}",
            ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml",
        )
        self.parts["[Content_Types].xml"] = ET.tostring(types)
        self.parts["xl/_rels/workbook.xml.rels"] = ET.tostring(self.rels)
        for view in self.book.iter(tag("workbookView")):
            view.set("activeTab", "0")
        self.parts["xl/workbook.xml"] = ET.tostring(self.book)
        self.refresh()
        return self.save()

    def mark_game(self, name: str, players: Sequence[str]) -> bytes | None:
        if self.has_game(name):
            return None
        lookup: dict[str, list[int]] = {}
        for row in range(2, self.last_player + 1):
            lookup.setdefault(normalized_name(str(self.value(f"A{row}"))), []).append(row)
        self.last_column += 1
        self.set_cell(1, self.last_column, name, style=self.style(1, 13))
        # Rebuild the four known summary rows below the extended player list.
        for xml_row in list(self.data):
            index = int(xml_row.attrib["r"])
            if index > self.last_player:
                self.data.remove(xml_row)
                self.rows.pop(index, None)
                for cell in xml_row:
                    self.cells.pop(cell.attrib["r"], None)
        for player in players:
            key = normalized_name(player)
            if key not in lookup:
                self.last_player += 1
                row = self.last_player
                lookup[key] = [row]
                for col in range(1, 14):
                    value: str | float | bool = player if col == 1 else False if col >= 4 else 0
                    self.set_cell(row, col, value, style=self.style(2, col))
            for row in lookup[key]:
                self.set_cell(row, self.last_column, 1, style=self.style(2, 14))
        self.refresh()
        return self.save()

    def refresh(self) -> None:
        last = self.last_player
        end = get_column_letter(self.last_column)
        totals = {row: self.total(row) for row in range(2, last + 1)}
        for row, total in totals.items():
            formula = f"C{row}+SUM(N{row}:{end}{row})" if self.last_column >= 14 else f"C{row}"
            self.set_cell(row, 2, total, formula=formula)
        summary = last + 1
        for offset, title in enumerate(("Всего", "Участники сезона", "Новички", "Активные игроки сезона")):
            self.set_cell(summary + offset, 1, title, style=self.style(1, 1))
        for col in (2, 3):
            letter = get_column_letter(col)
            count = sum(self.number(f"{letter}{row}") > 0 for row in totals)
            self.set_cell(
                summary, col, count, style=self.style(1, col), formula=f'COUNTIF({letter}2:{letter}{last},">0")'
            )
        for col in range(14, self.last_column + 1):
            letter = get_column_letter(col)
            attendance_count = sum(self.number(f"{letter}{row}", sum_cell=True) for row in totals)
            self.set_cell(
                summary, col, attendance_count, style=self.style(1, col), formula=f"SUM({letter}2:{letter}{last})"
            )
        for offset, threshold in ((1, 0), (3, 3)):
            count = sum(total > self.number(f"C{row}") + threshold for row, total in totals.items())
            self.set_cell(
                summary + offset,
                2,
                count,
                style=self.style(1, 2),
                formula=f"SUMPRODUCT(--(B2:B{last}>C2:C{last}+{threshold}))",
            )
        self.set_cell(
            summary + 2,
            2,
            self.number(f"B{summary}") - self.number(f"C{summary}"),
            style=self.style(1, 2),
            formula=f"B{summary}-C{summary}",
        )
        # Retain rule expressions and differential styles; extend to all players/games.
        for cf in list(self.sheet.findall(tag("conditionalFormatting"))):
            if any(rule.get("type") == "notContainsBlanks" for rule in cf):
                # The template limits this extra green rule to its first two
                # games. Extending it would also paint text such as ❌ green.
                continue
            old = cf.get("sqref", "")
            first = old.split(":", 1)[0]
            if re.fullmatch(r"A\d+", first):
                cf.set("sqref", f"A2:A{last}")
            elif re.fullmatch(r"B\d+", first):
                cf.set("sqref", f"B1:B{last}")
            elif self.last_column < 14:
                # Keep dormant formatting for the first future game without adding cells.
                cf.set("sqref", f"N2:N{last}")
                for cf_formula in cf.iter(tag("formula")):
                    cf_formula.text = re.sub(r"\b[A-Z]+2\b", "N2", cf_formula.text or "")
            else:
                cf.set("sqref", f"N2:{end}{last}")
                for cf_formula in cf.iter(tag("formula")):
                    cf_formula.text = re.sub(r"\b[A-Z]+2\b", "N2", cf_formula.text or "")
        if self.last_column >= 14 and not any(
            cf.get("sqref", "").startswith("N2:") for cf in self.sheet.findall(tag("conditionalFormatting"))
        ):
            # New seasons obtain the same green/yellow attendance rules as their source.
            source = ET.fromstring(self.parts[self._source_path()])  # noqa: S314
            for cf in source.findall(tag("conditionalFormatting")):
                if cf.get("sqref", "").startswith("N"):
                    clone = deepcopy(cf)
                    clone.set("sqref", f"N2:{end}{last}")
                    self._insert_cf(clone)
        dimension = self.sheet.find(tag("dimension"))
        if dimension is not None:
            dimension.set("ref", f"A1:{end}{summary + 3}")

    def _source_path(self) -> str:
        # Current sheet for a game update; original latest season for a new one.
        target = next(r.attrib["Target"] for r in self.rels if r.attrib["Id"] == self.season.attrib[f"{{{REL}}}id"])
        return target.lstrip("/") if target.startswith("/") else f"xl/{target}"

    def _insert_cf(self, rule: ET.Element) -> None:
        children = list(self.sheet)
        following = [i for i, c in enumerate(children) if c.tag in {tag("pageMargins"), tag("pageSetup")}]
        self.sheet.insert(min(following) if following else len(children), rule)

    def save(self) -> bytes:
        self.data[:] = sorted(self.data, key=lambda row: int(row.attrib["r"]))
        for row in self.data:
            row[:] = sorted(row, key=lambda c: column_index_from_string(re.sub(r"\d", "", c.attrib["r"])))
        self.parts[self.path] = ET.tostring(self.sheet, encoding="utf-8", xml_declaration=True)
        output = BytesIO()
        with ZipFile(output, "w", compression=ZIP_DEFLATED) as archive:
            for name, content in self.parts.items():
                archive.writestr(name, content)
        return output.getvalue()

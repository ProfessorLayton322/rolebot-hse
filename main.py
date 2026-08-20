import os
import asyncio
import logging
import requests
import openpyxl
import yadisk
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.utils.keyboard import ReplyKeyboardBuilder
from aiogram.client.session.aiohttp import AiohttpSession
from aiohttp import TCPConnector
import socket
import aiohttp

import socket

# --- НАСТРОЙКИ ---
TG_BOT_TOKEN = #
YANDEX_TOKEN = #

REGISTRATION_PATH = "registration.xlsx" # Таблица профилей
TABLE_PATH = "записи.xlsx"             # Таблица игр


dp = Dispatcher()
y = yadisk.YaDisk(token=YANDEX_TOKEN)

# --- СОСТОЯНИЯ (FSM) ---
class ProfileReg(StatesGroup):
    name = State()
    vk = State()
    crosspoly = State()
    experience = State()
    needs_pass = State()
    pass_cyrillic = State()
    pass_latin = State()
    pass_email = State()
    pass_citizen = State()

class GameReg(StatesGroup):
    wish_play = State()
    dont_wish_play = State()
    character_wish = State()

# --- КЛАВИАТУРЫ ---
def main_menu_kb():
    kb = ReplyKeyboardBuilder()
    kb.button(text="📝 Зарегистрировать профиль")
    kb.button(text="🎮 Записаться на игру")
    return kb.as_markup(resize_keyboard=True)

def yes_no_kb():
    kb = ReplyKeyboardBuilder()
    kb.button(text="Да")
    kb.button(text="Нет")
    return kb.as_markup(resize_keyboard=True)

# --- ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ДЛЯ EXCEL ---

def download_excel(disk_path):
    """Скачивает файл с Яндекс.Диска во временный локальный файл"""
    temp_file = f"temp_{os.path.basename(disk_path)}"
    if not y.is_file(disk_path):
        # Если файла нет на Диске, создаем пустой с заголовками
        wb = openpyxl.Workbook()
        ws = wb.active
        if disk_path == REGISTRATION_PATH:
            ws.append(["tg_id", "Фамилия Имя", "Ссылка ВК", "Кросспол", "Опыт", "ФИО Кириллица", "ФИО Латиница", "Email", "Гражданство РФ"])
        else:
            ws.append(["tg_id", "Фамилия Имя", "С кем играть", "С кем НЕ играть", "Пожелания по персонажу", "Статус прихода"])
        wb.save(temp_file)
        y.upload(temp_file, disk_path)
        return wb, temp_file

    download_url = y.get_download_link(disk_path)
    response = requests.get(download_url)
    with open(temp_file, "wb") as f:
        f.write(response.content)
    return openpyxl.load_workbook(temp_file), temp_file

def find_user_row_or_empty(sheet, tg_id, marker="№", column="A"):
    """
    Ищет строку с существующим tg_id для перезаписи. 
    Если не находит — возвращает первую пустую строку после маркера/данных.
    """
    # 1. Сначала ищем, нет ли уже такого юзера (проверка на уникальность / перезапись)
    for cell in sheet[column]:
        if str(cell.value) == str(tg_id):
            return cell.row

    # 2. Если юзера нет, ищем маркер для вставки в конец списка
    marker_row = None
    for cell in sheet[column]:
        if cell.value == marker:
            marker_row = cell.row
            break
            
    lower_bound = marker_row if marker_row else 1
    if marker_row:
        for merged_range in sheet.merged_cells.ranges:
            if marker_row in merged_range.rows:
                lower_bound = max(lower_bound, merged_range.max_row)
                break

    current_row = lower_bound + 1 if marker_row else sheet.max_row + 1
    while current_row <= sheet.max_row + 1:
        if sheet.cell(row=current_row, column=1).value is None:
            return current_row
        current_row += 1
    return current_row

def get_profile_data(tg_id):
    """Извлекает ФИО пользователя из таблицы регистраций для переноса в таблицу записей"""
    wb, temp_file = download_excel(REGISTRATION_PATH)
    sheet = wb.active
    name = "Неизвестно"
    for row in range(2, sheet.max_row + 1):
        if str(sheet.cell(row=row, column=1).value) == str(tg_id):
            name = sheet.cell(row=row, column=2).value
            break
    wb.close()
    if os.path.exists(temp_file): os.remove(temp_file)
    return name

# --- ХЭНДЛЕРЫ БОТА ---

@dp.message(Command("start"))
async def start_cmd(message: types.Message):
    await message.answer("Привет! Этот бот поможет тебе зарегистрироваться на LARP-игры.", reply_markup=main_menu_kb())

# --- ЦЕПОЧКА 1: РЕГИСТРАЦИЯ ПРОФИЛЯ ---

@dp.message(F.text == "📝 Зарегистрировать профиль")
async def reg_start(message: types.Message, state: FSMContext):
    await message.answer("Введите ваши Фамилию и Имя:")
    await state.set_state(ProfileReg.name)

@dp.message(ProfileReg.name)
async def reg_name(message: types.Message, state: FSMContext):
    await state.update_data(name=message.text)
    await message.answer("Введите ссылку на страницу ВК (или прочерк `-`):")
    await state.set_state(ProfileReg.vk)

@dp.message(ProfileReg.vk)
async def reg_vk(message: types.Message, state: FSMContext):
    await state.update_data(vk=message.text)
    await message.answer("Готовы ли вы кроссполить?", reply_markup=yes_no_kb())
    await state.set_state(ProfileReg.crosspoly)

@dp.message(ProfileReg.crosspoly)
async def reg_crosspoly(message: types.Message, state: FSMContext):
    await state.update_data(crosspoly=message.text)
    await message.answer("Играли ли вы в LARP до этого?", reply_markup=yes_no_kb())
    await state.set_state(ProfileReg.experience)

@dp.message(ProfileReg.experience)
async def reg_exp(message: types.Message, state: FSMContext):
    await state.update_data(experience=message.text)
    await message.answer("Нужен ли вам пропуск на локацию?", reply_markup=yes_no_kb())
    await state.set_state(ProfileReg.needs_pass)

@dp.message(ProfileReg.needs_pass)
async def reg_pass_choice(message: types.Message, state: FSMContext):
    if message.text.lower() == "да":
        await message.answer("Введите ФИО кириллицей (полностью):", reply_markup=types.ReplyKeyboardRemove())
        await state.set_state(ProfileReg.pass_cyrillic)
    else:
        # Если пропуск не нужен, забиваем поля прочерками и сохраняем
        data = await state.get_data()
        save_profile(message.from_user.id, data, has_pass=False)
        await message.answer("Профиль успешно сохранен!", reply_markup=main_menu_kb())
        await state.clear()

@dp.message(ProfileReg.pass_cyrillic)
async def reg_cyr(message: types.Message, state: FSMContext):
    await state.update_data(cyr=message.text)
    await message.answer("Введите ФИО латиницей (как в загранпаспорте):")
    await state.set_state(ProfileReg.pass_latin)

@dp.message(ProfileReg.pass_latin)
async def reg_lat(message: types.Message, state: FSMContext):
    await state.update_data(lat=message.text)
    await message.answer("Введите ваш email:")
    await state.set_state(ProfileReg.pass_email)

@dp.message(ProfileReg.pass_email)
async def reg_email(message: types.Message, state: FSMContext):
    await state.update_data(email=message.text)
    await message.answer("Являетесь ли вы гражданином РФ?", reply_markup=yes_no_kb())
    await state.set_state(ProfileReg.pass_citizen)

@dp.message(ProfileReg.pass_citizen)
async def reg_citizen(message: types.Message, state: FSMContext):
    await state.update_data(citizen=message.text)
    data = await state.get_data()
    save_profile(message.from_user.id, data, has_pass=True)
    await message.answer("Профиль с паспортными данными успешно сохранен!", reply_markup=main_menu_kb())
    await state.clear()

def save_profile(tg_id, data, has_pass=False):
    wb, temp_file = download_excel(REGISTRATION_PATH)
    sheet = wb.active
    row = find_user_row_or_empty(sheet, tg_id)
    
    row_data = [
        str(tg_id), data['name'], data['vk'], data['crosspoly'], data['experience'],
        data.get('cyr', '-') if has_pass else '-',
        data.get('lat', '-') if has_pass else '-',
        data.get('email', '-') if has_pass else '-',
        data.get('citizen', '-') if has_pass else '-'
    ]
    
    for col, val in enumerate(row_data, start=1):
        sheet.cell(row=row, column=col, value=val)
        
    wb.save(temp_file)
    y.upload(temp_file, REGISTRATION_PATH, overwrite=True)
    if os.path.exists(temp_file): os.remove(temp_file)

# --- ЦЕПОЧКА 2: ЗАПИСЬ НА КОНКРЕТНУЮ ИГРУ ---

@dp.message(F.text == "🎮 Записаться на игру")
async def game_start(message: types.Message, state: FSMContext):
    tg_id = message.from_user.id
    
    # Защита: проверяем регистрацию профиля
    wb, temp_file = download_excel(REGISTRATION_PATH)
    sheet = wb.active
    is_registered = any(str(cell.value) == str(tg_id) for cell in sheet["A"])
    wb.close()
    if os.path.exists(temp_file): os.remove(temp_file)

    if not is_registered:
        await message.answer("❌ Вы не можете записаться на игру, пока не зарегистрируете свой профиль! Пожалуйста, сначала нажмите кнопку '📝 Зарегистрировать профиль'.")
        return

    await message.answer("С кем бы вы ХОТЕЛИ играть?", reply_markup=types.ReplyKeyboardRemove())
    await state.set_state(GameReg.wish_play)

@dp.message(GameReg.wish_play)
async def game_wish(message: types.Message, state: FSMContext):
    await state.update_data(wish=message.text)
    await message.answer("С кем бы вы НЕ ХОТЕЛИ играть?")
    await state.set_state(GameReg.dont_wish_play)

@dp.message(GameReg.dont_wish_play)
async def game_dont_wish(message: types.Message, state: FSMContext):
    await state.update_data(dont_wish=message.text)
    await message.answer("Ваши пожелания по персонажу:")
    await state.set_state(GameReg.character_wish)

@dp.message(GameReg.character_wish)
async def game_final_save(message: types.Message, state: FSMContext):
    tg_id = message.from_user.id
    game_data = await state.get_data()
    
    # Получаем имя из таблицы профилей для кросс-ссылки данных
    user_name = get_profile_data(tg_id)
    
    wb, temp_file = download_excel(TABLE_PATH)
    sheet = wb.active
    row = find_user_row_or_empty(sheet, tg_id)
    
    row_data = [
        str(tg_id), user_name, game_data['wish'], game_data['dont_wish'], message.text, "Ожидается"
    ]
    
    for col, val in enumerate(row_data, start=1):
        sheet.cell(row=row, column=col, value=val)
        
    wb.save(temp_file)
    y.upload(temp_file, TABLE_PATH, overwrite=True)
    if os.path.exists(temp_file): os.remove(temp_file)
    
    await message.answer("🎲 Ваша заявка на игру принята или обновлена!", reply_markup=main_menu_kb())
    await state.clear()

# --- ПОЛЛИНГ ---
async def main():
    logging.basicConfig(level=logging.INFO)
    bot = Bot(token=TG_BOT_TOKEN)
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())